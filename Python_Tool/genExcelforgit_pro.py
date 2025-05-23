import git
import os # csv 模組不再直接使用於檔案生成，但 os 可能仍用於路徑處理
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta, date
import openpyxl # 匯入 openpyxl
from openpyxl.styles import Font, Alignment, PatternFill # 用於設定儲存格樣式

def generate_git_log_excel(repo_path, output_excel_file, num_commits=None, start_date_str=None, end_date_str=None):
    """
    生成一份 Excel 檔案，其中包含 Git 倉儲的提交歷史。
    最舊的 commit 在最上面，並在日期變更時插入日期分隔列。

    參數：
        repo_path (str): Git 倉儲的路徑。
        output_excel_file (str): 輸出的 Excel 檔案名稱。
        num_commits (int, optional): 要擷取的提交筆數。如果為 None，則擷取所有符合條件的提交。
        start_date_str (str, optional): 開始日期字串 (YYYY-MM-DD)。
        end_date_str (str, optional): 結束日期字串 (YYYY-MM-DD)。
    """
    try:
        repo = git.Repo(repo_path)
    except git.exc.InvalidGitRepositoryError:
        messagebox.showerror("錯誤", f"指定的路徑 '{repo_path}' 不是一個有效的 Git 倉儲。")
        return False
    except git.exc.NoSuchPathError:
        messagebox.showerror("錯誤", f"指定的路徑 '{repo_path}' 不存在。")
        return False
    except Exception as e:
        messagebox.showerror("錯誤", f"讀取倉儲 '{repo_path}' 時發生未知錯誤：{e}")
        return False

    iter_kwargs = {}
    try:
        if start_date_str:
            dt_start = datetime.strptime(start_date_str, "%Y-%m-%d")
            iter_kwargs['since'] = dt_start
        if end_date_str:
            dt_end = datetime.strptime(end_date_str, "%Y-%m-%d")
            iter_kwargs['until'] = dt_end + timedelta(days=1)
    except ValueError:
        messagebox.showerror("日期格式錯誤", "內部日期處理錯誤，格式應為YYYY-MM-DD。")
        return False
    except Exception as e:
        messagebox.showerror("日期處理錯誤", f"處理日期時發生錯誤: {e}")
        return False

    headers = ["姓名", "SHA值", "commit內容"]
    # 儲存包含日期的提交詳細資料: [commit_date_obj, author_name, sha, commit_message]
    all_commit_details_with_date = []

    try:
        commits_iterator = repo.iter_commits(**iter_kwargs)
        for commit in commits_iterator:
            commit_date = commit.committed_datetime.date() # 獲取 commit 的日期物件
            author_name = commit.author.name
            sha = commit.hexsha
            commit_message = commit.message.strip().replace('\n', ' ') # 清理換行符
            all_commit_details_with_date.append([commit_date, author_name, sha, commit_message])

        all_commit_details_with_date.reverse() # 最舊的在前

        if num_commits and num_commits > 0:
            final_commits_to_process = all_commit_details_with_date[:num_commits]
        else:
            final_commits_to_process = all_commit_details_with_date

        if not final_commits_to_process and (start_date_str or end_date_str):
            messagebox.showinfo("提示", "在指定的日期範圍內沒有找到任何 commit。")
            # 即使沒有 commit，也可能需要產生一個空的 Excel 檔案或只有表頭的檔案

    except Exception as e:
        messagebox.showerror("錯誤", f"處理 Git 提交時發生錯誤：{e}")
        return False

    # --- 開始寫入 Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Git提交紀錄"

    # 寫入表頭
    ws.append(headers)
    # 設定表頭樣式
    for col_idx, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    current_processing_date = None
    if not final_commits_to_process and not (start_date_str or end_date_str) and not repo.head.is_valid(): # 處理完全空的倉儲
         messagebox.showinfo("提示", "此倉儲中沒有任何 commit。")


    for commit_item in final_commits_to_process:
        commit_date_obj = commit_item[0]
        actual_commit_data = commit_item[1:] # [author_name, sha, commit_message]

        if commit_date_obj != current_processing_date:
            current_processing_date = commit_date_obj
            # 插入日期分隔列
            date_str_display = current_processing_date.strftime("%Y-%m-%d")
            # 建立一個只包含日期的列，其他儲存格為空以便合併
            ws.append([date_str_display] + [''] * (len(headers) - 1))
            
            # 合併儲存格並設定樣式
            date_row_idx = ws.max_row # 獲取剛插入的日期列的索引
            ws.merge_cells(start_row=date_row_idx, start_column=1, end_row=date_row_idx, end_column=len(headers))
            merged_cell = ws.cell(row=date_row_idx, column=1)
            merged_cell.font = Font(bold=True, color="FFFFFF") # 白色字體
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            merged_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # 藍色背景

        # 寫入實際的 commit 資料
        ws.append(actual_commit_data)

    # 自動調整欄寬 (可選)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width < 70 else 70 # 限制最大寬度


    try:
        wb.save(output_excel_file)
        return True
    except Exception as e:
        messagebox.showerror("錯誤", f"儲存 Excel 檔案 '{output_excel_file}' 時發生錯誤：{e}")
        return False

class GitLogGUI:
    def __init__(self, master):
        self.master = master
        master.title("Git 提交紀錄轉 Excel 工具") # 更新視窗標題
        master.geometry("600x400")

        style = ttk.Style()
        style.configure("TLabel", padding=5, font=('Arial', 10))
        style.configure("TButton", padding=5, font=('Arial', 10))
        style.configure("TEntry", padding=5, font=('Arial', 10))

        self.repo_path_label = ttk.Label(master, text="Git 倉儲路徑:")
        self.repo_path_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.repo_path_var = tk.StringVar()
        self.repo_path_var.trace_add("write", lambda *args: self.update_default_output_filename())
        self.repo_path_entry = ttk.Entry(master, textvariable=self.repo_path_var, width=45)
        self.repo_path_entry.grid(row=0, column=1, columnspan=2, padx=10, pady=5, sticky="ew")
        self.browse_repo_button = ttk.Button(master, text="瀏覽...", command=self.browse_repo)
        self.browse_repo_button.grid(row=0, column=3, padx=10, pady=5)

        self.output_file_label = ttk.Label(master, text="輸出 Excel 檔案:") # 更新標籤
        self.output_file_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        
        initial_today_str_for_filename = date.today().strftime("%Y%m%d")
        self.output_file_var = tk.StringVar(value=f"{initial_today_str_for_filename}_git_commit_history.xlsx") # 更新預設副檔名
        
        self.output_file_entry = ttk.Entry(master, textvariable=self.output_file_var, width=45)
        self.output_file_entry.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky="ew")
        self.save_as_button = ttk.Button(master, text="另存為...", command=self.save_as_output)
        self.save_as_button.grid(row=1, column=3, padx=10, pady=5)

        today_str_for_date_fields = date.today().strftime("%Y-%m-%d")

        self.start_date_label = ttk.Label(master, text="開始日期 (YYYY-MM-DD):")
        self.start_date_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.start_date_var = tk.StringVar(value=today_str_for_date_fields)
        self.start_date_entry = ttk.Entry(master, textvariable=self.start_date_var, width=20)
        self.start_date_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")

        self.end_date_label = ttk.Label(master, text="結束日期 (YYYY-MM-DD):")
        self.end_date_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.end_date_var = tk.StringVar(value=today_str_for_date_fields)
        self.end_date_entry = ttk.Entry(master, textvariable=self.end_date_var, width=20)
        self.end_date_entry.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        self.num_commits_label = ttk.Label(master, text="擷取筆數 (0或空=全部):")
        self.num_commits_label.grid(row=4, column=0, padx=10, pady=5, sticky="w")
        self.num_commits_var = tk.StringVar(value="0")
        self.num_commits_entry = ttk.Entry(master, textvariable=self.num_commits_var, width=10)
        self.num_commits_entry.grid(row=4, column=1, padx=10, pady=5, sticky="w")

        self.generate_button = ttk.Button(master, text="生成 Excel", command=self.generate) # 更新按鈕文字
        self.generate_button.grid(row=5, column=0, columnspan=4, padx=10, pady=15)

        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(master, textvariable=self.status_var, font=('Arial', 10, 'italic'))
        self.status_label.grid(row=6, column=0, columnspan=4, padx=10, pady=5, sticky="w")

        master.grid_columnconfigure(1, weight=1)

    def update_default_output_filename(self):
        repo_path = self.repo_path_var.get()
        today_str_filename = date.today().strftime("%Y%m%d")
        
        if not repo_path:
            self.output_file_var.set(f"{today_str_filename}_git_commit_history.xlsx")
            return

        try:
            repo = git.Repo(repo_path)
            latest_commit = None
            try:
                latest_commit = next(repo.iter_commits(max_count=1))
            except StopIteration:
                pass

            if latest_commit:
                latest_sha_short = latest_commit.hexsha[:7]
                default_filename = f"{today_str_filename}_{latest_sha_short}_git_commit_history.xlsx" # 更新副檔名
            else:
                default_filename = f"{today_str_filename}_no_commits_history.xlsx" # 更新副檔名
            
            self.output_file_var.set(default_filename)

        except (git.exc.InvalidGitRepositoryError, git.exc.NoSuchPathError):
            self.output_file_var.set(f"{today_str_filename}_git_commit_history.xlsx") # 更新副檔名
        except Exception:
            self.output_file_var.set(f"{today_str_filename}_git_commit_history.xlsx") # 更新副檔名


    def browse_repo(self):
        path = filedialog.askdirectory(title="選擇 Git 倉儲資料夾")
        if path:
            self.repo_path_var.set(path)
            self.status_var.set(f"已選擇倉儲: {path}")

    def save_as_output(self):
        self.update_default_output_filename()
        
        path = filedialog.asksaveasfilename(
            title="儲存 Excel 檔案", # 更新標題
            initialfile=self.output_file_var.get(),
            defaultextension=".xlsx", # 更新預設副檔名
            filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")] # 更新檔案類型
        )
        if path:
            self.output_file_var.set(path)
            self.status_var.set(f"Excel 將儲存至: {path}") # 更新訊息

    def generate(self):
        repo_path = self.repo_path_var.get()
        output_file = self.output_file_var.get()
        num_commits_str = self.num_commits_var.get()
        start_date_input = self.start_date_var.get().strip()
        end_date_input = self.end_date_var.get().strip()

        if not repo_path:
            messagebox.showwarning("輸入錯誤", "請選擇 Git 倉儲路徑。")
            return
        if not output_file:
            messagebox.showwarning("輸入錯誤", "請指定輸出的 Excel 檔案名稱或路徑。")
            return
        if not output_file.lower().endswith(".xlsx"): # 簡單檢查副檔名
             output_file += ".xlsx"
             self.output_file_var.set(output_file) # 更新UI上的檔名
             messagebox.showinfo("檔名提示", f"檔名已自動添加 .xlsx 副檔名。\n將儲存為: {output_file}")


        num_commits = 0
        if num_commits_str:
            try:
                num_commits = int(num_commits_str)
                if num_commits < 0:
                    messagebox.showwarning("輸入錯誤", "擷取筆數不能為負數。")
                    return
            except ValueError:
                messagebox.showwarning("輸入錯誤", "擷取筆數必須是一個有效的數字。")
                return

        if start_date_input:
            try:
                datetime.strptime(start_date_input, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("日期格式錯誤", "開始日期格式必須為YYYY-MM-DD。")
                return
        
        if end_date_input:
            try:
                datetime.strptime(end_date_input, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("日期格式錯誤", "結束日期格式必須為YYYY-MM-DD。")
                return
        
        if start_date_input and end_date_input:
            try:
                dt_start_check = datetime.strptime(start_date_input, "%Y-%m-%d").date()
                dt_end_check = datetime.strptime(end_date_input, "%Y-%m-%d").date()
                if dt_start_check > dt_end_check:
                    messagebox.showwarning("日期錯誤", "開始日期不能晚於結束日期。")
                    return
            except ValueError:
                return

        self.status_var.set("處理中，請稍候...")
        self.master.update_idletasks()

        # 呼叫新的 Excel 生成函式
        success = generate_git_log_excel(
            repo_path,
            output_file,
            num_commits if num_commits > 0 else None,
            start_date_input if start_date_input else None,
            end_date_input if end_date_input else None
        )

        if success:
            self.status_var.set(f"成功生成 Excel 檔案: '{os.path.abspath(output_file)}'")
            messagebox.showinfo("成功", f"Excel 檔案已成功生成！\n儲存於: {os.path.abspath(output_file)}")
        else:
            self.status_var.set("生成失敗或過程中出現問題。")


if __name__ == "__main__":
    root = tk.Tk()
    gui = GitLogGUI(root)
    root.mainloop()
