import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime
import configparser
import re

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class DateInputDialog(tk.Toplevel):
    # --- 修改點 1：接收傳入的預設日期 ---
    def __init__(self, parent, default_d1, default_d2):
        super().__init__(parent)
        self.title("設定比對日期")
        self.geometry("320x180")
        self.result = None
        
        tk.Label(self, text="請輸入 Bin File 1 目標日期 (較舊)\n(格式 YYMMDD，例如: 260315):", pady=5).pack()
        self.date1_entry = tk.Entry(self, justify="center")
        self.date1_entry.pack()
        
        tk.Label(self, text="請輸入 Bin File 2 目標日期 (較新)\n(格式 YYMMDD，例如: 260316):", pady=5).pack()
        self.date2_entry = tk.Entry(self, justify="center")
        self.date2_entry.pack()
        
        # 填入自動計算出的預設日期
        self.date1_entry.insert(0, default_d1)
        self.date2_entry.insert(0, default_d2)
        
        tk.Button(self, text="確定搜尋", command=self.on_ok, bg="#4CAF50", fg="white", width=15).pack(pady=10)
        
        self.transient(parent)
        self.grab_set()
        self.wait_window(self)
        
    def on_ok(self):
        d1 = self.date1_entry.get().strip()
        d2 = self.date2_entry.get().strip()
        if d1 and d2:
            self.result = (d1, d2)
            self.destroy()
        else:
            messagebox.showwarning("警告", "請輸入兩個日期！", parent=self)

class BinComparatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Bin File 比較工具 - 智慧日期預填版")
        self.root.geometry("800x600") 
        
        self.base_bin_dir = 'Bin_Files'
        self.sub_dirs = ['InfoBlock', 'IDPage', 'systeminfo']
        self.struct_dir = 'Struct'
        self.result_dir = 'Result'
        self.config_file = 'config.ini'
        
        self.setup_directories()
        self.setup_config()
        
        self.current_category = tk.StringVar(value=self.sub_dirs[0])
        self.bin1_path = tk.StringVar()
        self.bin2_path = tk.StringVar()
        self.struct_path = tk.StringVar()
        
        self.compare_results = []       
        self.used_files = {}            
        self.comparison_stats = {}      

        self.create_widgets()

    def setup_directories(self):
        for dir_name in [self.base_bin_dir, self.struct_dir, self.result_dir]:
            if not os.path.exists(dir_name):
                os.makedirs(dir_name)
                
        for sub in self.sub_dirs:
            sub_path = os.path.join(self.base_bin_dir, sub)
            if not os.path.exists(sub_path):
                os.makedirs(sub_path)

    def setup_config(self):
        if not os.path.exists(self.config_file):
            with open(self.config_file, 'w', encoding='utf-8') as f:
                f.write("[Blacklist]\n")
                f.write("# 設定略過比對的 Byte 範圍。支援格式: 123~234 或 Byte_123~Byte_234\n")
                f.write("# 若有多組範圍，請用逗號分隔。例如: Byte_10~Byte_20, 100~200\n")
                for sub in self.sub_dirs:
                    f.write(f"{sub} = \n")

    def get_blacklist(self):
        blacklist = {sub: [] for sub in self.sub_dirs}
        config = configparser.ConfigParser()
        
        if os.path.exists(self.config_file):
            config.read(self.config_file, encoding='utf-8')
            if 'Blacklist' in config:
                for sub in self.sub_dirs:
                    raw_str = config['Blacklist'].get(sub, '')
                    if not raw_str.strip():
                        continue
                    
                    parts = raw_str.split(',')
                    for p in parts:
                        nums = re.findall(r'\d+', p)
                        if len(nums) >= 2:
                            start_bl = int(nums[0])
                            end_bl = int(nums[1])
                            blacklist[sub].append((min(start_bl, end_bl), max(start_bl, end_bl)))
        return blacklist

    def create_widgets(self):
        frame_files = tk.LabelFrame(self.root, text="手動單次比較設定", padx=10, pady=10)
        frame_files.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_files, text="選擇比較項目:").grid(row=0, column=0, sticky="e")
        category_cb = ttk.Combobox(frame_files, textvariable=self.current_category, values=self.sub_dirs, state="readonly", width=47)
        category_cb.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        category_cb.bind('<<ComboboxSelected>>', self.on_category_change)

        tk.Label(frame_files, text="Bin File 1:").grid(row=1, column=0, sticky="e")
        tk.Entry(frame_files, textvariable=self.bin1_path, width=50).grid(row=1, column=1, padx=5, pady=2)
        tk.Button(frame_files, text="瀏覽", command=lambda: self.select_bin_file(self.bin1_path)).grid(row=1, column=2)

        tk.Label(frame_files, text="Bin File 2:").grid(row=2, column=0, sticky="e")
        tk.Entry(frame_files, textvariable=self.bin2_path, width=50).grid(row=2, column=1, padx=5, pady=2)
        tk.Button(frame_files, text="瀏覽", command=lambda: self.select_bin_file(self.bin2_path)).grid(row=2, column=2)

        tk.Label(frame_files, text="Struct (Excel):").grid(row=3, column=0, sticky="e")
        tk.Entry(frame_files, textvariable=self.struct_path, width=50).grid(row=3, column=1, padx=5, pady=2)
        tk.Button(frame_files, text="瀏覽", command=lambda: self.select_file(self.struct_path, self.struct_dir, [("Excel files", "*.xlsx *.xls")])).grid(row=3, column=2)

        frame_btns = tk.Frame(self.root)
        frame_btns.pack(fill="x", padx=10, pady=5)
        
        tk.Button(frame_btns, text="🚀 一鍵智慧日期比對 (含自動匯出)", command=self.auto_compare_all, bg="#FF9800", fg="white", font=("Arial", 11, "bold")).pack(side="left", padx=5)
        tk.Button(frame_btns, text="單次手動比較", command=self.compare_files_manual, bg="#4CAF50", fg="white", font=("Arial", 10)).pack(side="left", padx=5)
        tk.Button(frame_btns, text="清空列表", command=self.clear_results, font=("Arial", 10)).pack(side="left", padx=5)
        tk.Button(frame_btns, text="手動匯出 Excel", command=self.export_excel, bg="#2196F3", fg="white", font=("Arial", 10, "bold")).pack(side="right", padx=5)

        frame_result = tk.LabelFrame(self.root, text="不一致結果 (逐 Byte 列出)", padx=10, pady=10)
        frame_result.pack(fill="both", expand=True, padx=10, pady=5)

        columns = ("Category", "Variable", "Diff Byte Addr", "Bin 1 Value", "Bin 2 Value")
        self.tree = ttk.Treeview(frame_result, columns=columns, show="headings")
        
        col_headings = {"Category": "Category", "Variable": "Variable", "Diff Byte Addr": "差異 Byte 位置", "Bin 1 Value": "Bin 1 Value", "Bin 2 Value": "Bin 2 Value"}
        col_widths = {"Category": 80, "Variable": 120, "Diff Byte Addr": 100, "Bin 1 Value": 100, "Bin 2 Value": 100}
        
        for col in columns:
            self.tree.heading(col, text=col_headings[col])
            self.tree.column(col, width=col_widths[col], anchor="center")
        
        scrollbar = ttk.Scrollbar(frame_result, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

    def on_category_change(self, event):
        self.bin1_path.set("")
        self.bin2_path.set("")

    def select_bin_file(self, var):
        current_sub_dir = self.current_category.get()
        target_dir = os.path.join(self.base_bin_dir, current_sub_dir)
        self.select_file(var, target_dir, [("Bin files", "*.bin"), ("All files", "*.*")])

    def select_file(self, var, initialdir, filetypes=[("All files", "*.*")]):
        filepath = filedialog.askopenfilename(initialdir=initialdir, title="選擇檔案", filetypes=filetypes)
        if filepath:
            var.set(filepath)

    def clear_results(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.compare_results.clear()
        self.used_files.clear()
        self.comparison_stats.clear()

    # --- 新增點 2：掃描資料夾，找出最新的兩個日期 ---
    def get_latest_two_dates(self):
        unique_dates = set()
        
        for sub_dir in self.sub_dirs:
            bin_folder = os.path.join(self.base_bin_dir, sub_dir)
            if not os.path.exists(bin_folder):
                continue
                
            target_suffix = f"{sub_dir.lower()}.bin"
            try:
                for f in os.listdir(bin_folder):
                    if f.lower().endswith(target_suffix):
                        parts = f.split('_')
                        if len(parts) >= 1:
                            date_str = parts[0]
                            # 驗證是否為合法日期
                            try:
                                datetime.strptime(date_str, "%y%m%d")
                                unique_dates.add(date_str)
                            except ValueError:
                                pass
            except Exception:
                pass
                
        # 排序日期 (由大到小，即由新到舊)
        sorted_dates = sorted(list(unique_dates), reverse=True)
        
        if len(sorted_dates) >= 2:
            # 回傳：較舊的(次新), 較新的(最新)
            return sorted_dates[1], sorted_dates[0]
        elif len(sorted_dates) == 1:
            return sorted_dates[0], sorted_dates[0]
        else:
            # 如果都沒找到檔案，給個預設值
            return "260315", "260316"

    def find_closest_file(self, file_list, target_dt):
        closest_file = None
        min_diff = None
        for f in file_list:
            parts = f.split('_')
            if len(parts) >= 1:
                date_str = parts[0]
                try:
                    file_dt = datetime.strptime(date_str, "%y%m%d")
                    diff_days = abs((file_dt - target_dt).days)
                    if min_diff is None or diff_days < min_diff:
                        min_diff = diff_days
                        closest_file = f
                except ValueError:
                    continue
        return closest_file

    def auto_compare_all(self):
        # --- 修改點 3：取得最新日期並傳給 Dialog ---
        d1_default, d2_default = self.get_latest_two_dates()
        dialog = DateInputDialog(self.root, d1_default, d2_default)
        
        if not dialog.result:
            return
            
        target_date1_str, target_date2_str = dialog.result
        try:
            target_dt1 = datetime.strptime(target_date1_str, "%y%m%d")
            target_dt2 = datetime.strptime(target_date2_str, "%y%m%d")
        except ValueError:
            messagebox.showerror("錯誤", "日期格式不正確！")
            return

        self.clear_results()
        blacklist = self.get_blacklist()
        total_mismatch_vars = 0
        error_msgs = []
        file_log = []

        for sub_dir in self.sub_dirs:
            bin_folder = os.path.join(self.base_bin_dir, sub_dir)
            struct_file = os.path.join(self.struct_dir, f"{sub_dir}.xlsx")

            if not os.path.exists(struct_file):
                error_msgs.append(f"[{sub_dir}] 略過：找不到 {sub_dir}.xlsx")
                continue

            target_suffix = f"{sub_dir.lower()}.bin"
            bin_files = [f for f in os.listdir(bin_folder) if f.lower().endswith(target_suffix)]
            
            if len(bin_files) < 2:
                error_msgs.append(f"[{sub_dir}] 略過：找不到足夠符合 {target_suffix} 結尾的 Bin 檔案")
                continue

            file1 = self.find_closest_file(bin_files, target_dt1)
            file2 = self.find_closest_file(bin_files, target_dt2)

            if not file1 or not file2:
                error_msgs.append(f"[{sub_dir}] 略過：無法解析出有效日期")
                continue
            if file1 == file2:
                error_msgs.append(f"[{sub_dir}] 警告：兩個日期皆對應到同個檔案")
                continue

            file_log.append(f"[{sub_dir}] Bin1: {file1} | Bin2: {file2}")

            bin1_path = os.path.join(bin_folder, file1)
            bin2_path = os.path.join(bin_folder, file2)

            try:
                mismatch_vars = self._compare_core(bin1_path, bin2_path, struct_file, sub_dir, blacklist[sub_dir])
                total_mismatch_vars += mismatch_vars
            except Exception as e:
                error_msgs.append(f"[{sub_dir}] 發生錯誤：{str(e)}")

        exported_file_path = self.export_excel(show_popup=False)

        report_msg = f"智慧掃描比對完成！\n共發現 {total_mismatch_vars} 個變數含有不一致情形。\n"
        if exported_file_path:
            report_msg += f"\n📁 報表已自動匯出至：\n{exported_file_path}\n"
        
        report_msg += "\n--- 實際取用檔案紀錄 ---\n" + "\n".join(file_log)
        
        if error_msgs:
            report_msg += "\n\n--- ⚠️ 注意事項 ---\n" + "\n".join(error_msgs)
            messagebox.showwarning("完成帶有警告", report_msg)
        else:
            messagebox.showinfo("全部完成", report_msg)

    def compare_files_manual(self):
        bin1 = self.bin1_path.get()
        bin2 = self.bin2_path.get()
        struct_file = self.struct_path.get()
        current_cat = self.current_category.get()

        if not all([bin1, bin2, struct_file]):
            messagebox.showwarning("警告", "請先選擇兩個 Bin 檔案與 Struct 檔案！")
            return

        blacklist = self.get_blacklist()
        
        try:
            mismatch_vars = self._compare_core(bin1, bin2, struct_file, current_cat, blacklist[current_cat])
            if mismatch_vars == 0:
                messagebox.showinfo("完成", f"[{current_cat}] 區塊比對完成！內容完全一致。")
            else:
                messagebox.showinfo("完成", f"[{current_cat}] 區塊比對完成！發現 {mismatch_vars} 個變數含有不一致情形。")
        except Exception as e:
            messagebox.showerror("錯誤", f"發生錯誤：\n{str(e)}")

    def _compare_core(self, bin1_path, bin2_path, struct_file, category, current_blacklist):
        df_struct = pd.read_excel(struct_file, header=None)
        
        with open(bin1_path, 'rb') as f1, open(bin2_path, 'rb') as f2:
            data1 = f1.read()
            data2 = f2.read()

        filename1 = os.path.basename(bin1_path)
        filename2 = os.path.basename(bin2_path)
        self.used_files[category] = {'bin1': filename1, 'bin2': filename2}

        total_vars = 0
        mismatch_var_count = 0 
        
        for index, row in df_struct.iterrows():
            try:
                start_byte = int(row.iloc[1])
                length = int(row.iloc[2])
            except ValueError:
                continue 

            var_name = str(row.iloc[0])
            var_end_byte = start_byte + length - 1 

            is_blacklisted = False
            for bl_start, bl_end in current_blacklist:
                if max(start_byte, bl_start) <= min(var_end_byte, bl_end):
                    is_blacklisted = True
                    break
                    
            if is_blacklisted:
                continue 
            
            total_vars += 1 

            if start_byte >= len(data1) and start_byte >= len(data2):
                continue

            val1 = data1[start_byte : start_byte + length]
            val2 = data2[start_byte : start_byte + length]

            if val1 != val2:
                mismatch_var_count += 1
                
                for i in range(length):
                    b1 = val1[i] if i < len(val1) else None
                    b2 = val2[i] if i < len(val2) else None
                    
                    if b1 != b2:
                        diff_addr = start_byte + i 
                        hex_val1 = f"0x{b1:02X}" if b1 is not None else "N/A"
                        hex_val2 = f"0x{b2:02X}" if b2 is not None else "N/A"

                        record = (category, var_name, diff_addr, hex_val1, hex_val2)
                        self.compare_results.append(record)
                        self.tree.insert("", "end", values=record)
                
        self.comparison_stats[category] = {'total': total_vars, 'diffs': mismatch_var_count}
        return mismatch_var_count

    def export_excel(self, show_popup=True):
        if not self.comparison_stats:
            if show_popup:
                messagebox.showwarning("警告", "請先執行比對後再匯出報表！")
            return None

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_path = os.path.join(self.result_dir, f"Comparison_Result_{timestamp}.xlsx")
            
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                summary_data = []
                for cat in self.sub_dirs:
                    if cat in self.comparison_stats:
                        stats = self.comparison_stats[cat]
                        total = stats['total']
                        diffs = stats['diffs'] 
                        pct = (diffs / total) if total > 0 else 0
                        summary_data.append([cat, total, diffs, pct])
                    else:
                        summary_data.append([cat, 0, 0, 0])
                
                df_summary = pd.DataFrame(summary_data, columns=["分析區塊", "總變數數", "不一致變數數量", "不一致百分比"])
                df_summary.to_excel(writer, sheet_name="summary", index=False)
                
                for cat in self.sub_dirs:
                    cat_records = [rec for rec in self.compare_results if rec[0] == cat]
                    
                    if cat in self.used_files:
                        f1_name = self.used_files[cat]['bin1']
                        f2_name = self.used_files[cat]['bin2']
                    else:
                        f1_name = "Bin_File_1"
                        f2_name = "Bin_File_2"
                        
                    col_f1 = f"{f1_name}_數值"
                    col_f2 = f"{f2_name}_數值"
                    
                    sheet_data = [[rec[1], rec[2], rec[3], rec[4]] for rec in cat_records]
                    df_cat = pd.DataFrame(sheet_data, columns=["變數名稱", "差異 Byte 位置", col_f1, col_f2])
                    df_cat.to_excel(writer, sheet_name=cat, index=False)
            
                workbook = writer.book
                
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                align_center = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    
                    ws.auto_filter.ref = ws.dimensions
                    
                    for col in ws.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        for cell in col:
                            cell.alignment = align_center
                            cell.border = thin_border
                            
                            if cell.row == 1:
                                cell.fill = header_fill
                                cell.font = header_font
                            elif cell.row % 2 == 0:
                                cell.fill = alt_fill
                                
                            if sheet_name == "summary" and cell.column == 4 and cell.row > 1:
                                cell.number_format = '0.00%'

                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        ws.column_dimensions[column_letter].width = max_length + 6

            if show_popup:
                messagebox.showinfo("匯出成功", f"多工作表報表已成功儲存至：\n{export_path}")
            return export_path

        except Exception as e:
            if show_popup:
                messagebox.showerror("錯誤", f"匯出失敗：\n{str(e)}")
            return None

if __name__ == "__main__":
    root = tk.Tk()
    app = BinComparatorApp(root)
    root.mainloop()
